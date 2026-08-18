from decimal import Decimal
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.db import get_conn


INVOICE_REVIEW_COLUMNS = [
    "Rechnungs-ID",
    "Rechnungsnummer",
    "Dateiname",
    "Kunde/Lieferant",
    "Strasse",
    "Hausnummer",
    "PLZ",
    "Stadt",
    "Dokumenttyp",
    "Belegdatum",
    "Netto",
    "USt",
    "Brutto",
    "Validierungsstatus",
    "Validierungsfehler",
]

INVOICE_POS_REVIEW_COLUMNS = [
    "Rechnungs-ID",
    "Rechnungsnummer",
    "Kunde/Lieferant",
    "Belegdatum",
    "Position",
    "Positions-Netto",
    "Positions-Brutto",
]


def export_invoice_review_to_excel() -> BytesIO:
    with get_conn() as conn:
        invoice_rows = conn.execute(_invoice_review_query()).fetchall()
        pos_rows = conn.execute(_invoice_review_pos_query()).fetchall()

    positions_by_invoice_id: dict[int, list[dict[str, Any]]] = {}
    for pos in pos_rows:
        positions_by_invoice_id.setdefault(pos["Rechnungs-ID"], []).append(pos)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "invoice_review"
    worksheet.freeze_panes = "A2"

    header_font = Font(bold=True, color="000000")
    nested_header_font = Font(bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="F7F7F7")
    invoice_row_fill = PatternFill("solid", fgColor="FFFF00")
    pos_fill = PatternFill("solid", fgColor="F7F7F7")
    thick_side = Side(style="medium", color="000000")
    thin_side = Side(style="thin", color="D9D9D9")
    header_border = Border(top=thick_side, right=thick_side, bottom=thick_side, left=thick_side)
    data_border = Border(top=thin_side, right=thin_side, bottom=thin_side, left=thin_side)
    right_alignment = Alignment(horizontal="right", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")

    current_row = 1
    _write_row(
        worksheet,
        current_row,
        1,
        INVOICE_REVIEW_COLUMNS,
        font=header_font,
        fill=header_fill,
        border=header_border,
        alignment=left_alignment,
    )
    current_row += 1

    for invoice in invoice_rows:
        invoice_values = [_format_export_value(invoice.get(column), column) for column in INVOICE_REVIEW_COLUMNS]
        _write_row(
            worksheet,
            current_row,
            1,
            invoice_values,
            fill=invoice_row_fill,
            border=data_border,
            alignment=left_alignment,
        )
        for column_index in (1, 11, 12, 13, 15):
            worksheet.cell(row=current_row, column=column_index).alignment = right_alignment
        current_row += 1

        positions = positions_by_invoice_id.get(invoice["Rechnungs-ID"], [])
        if positions:
            _write_row(
                worksheet,
                current_row,
                2,
                INVOICE_POS_REVIEW_COLUMNS,
                font=nested_header_font,
                fill=pos_fill,
                border=header_border,
                alignment=left_alignment,
            )
            current_row += 1

            for pos in positions:
                pos_values = [_format_export_value(pos.get(column), column) for column in INVOICE_POS_REVIEW_COLUMNS]
                _write_row(worksheet, current_row, 2, pos_values, border=data_border, alignment=left_alignment)
                for column_index in (2, 6, 7, 8):
                    worksheet.cell(row=current_row, column=column_index).alignment = right_alignment
                current_row += 1

    _fit_columns(worksheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _invoice_review_query() -> str:
    return """
        SELECT
            i.id AS "Rechnungs-ID",
            i.invoice_number AS "Rechnungsnummer",
            d.file_name AS "Dateiname",
            c.name_original AS "Kunde/Lieferant",
            c.street AS "Strasse",
            c.house_number AS "Hausnummer",
            c.postal_code AS "PLZ",
            c.city AS "Stadt",
            i.invoice_type AS "Dokumenttyp",
            i.invoice_date AS "Belegdatum",
            i.gesamt_netto AS "Netto",
            i.tva AS "USt",
            i.gesamtbetrag AS "Brutto",
            CASE
                WHEN COALESCE(v.failed_count, 0) = 0 THEN 'ok'
                ELSE 'review_required'
            END AS "Validierungsstatus",
            COALESCE(v.failed_count, 0) AS "Validierungsfehler"
        FROM invoices i
        JOIN clients c ON c.id = i.client_id
        JOIN documents d ON d.id = i.document_id
        LEFT JOIN (
            SELECT invoice_id, COUNT(*) FILTER (WHERE NOT passed) AS failed_count
            FROM validation_results
            GROUP BY invoice_id
        ) v ON v.invoice_id = i.id
        ORDER BY i.id
    """


def _invoice_review_pos_query() -> str:
    return """
        SELECT
            i.id AS "Rechnungs-ID",
            i.invoice_number AS "Rechnungsnummer",
            c.name_original AS "Kunde/Lieferant",
            i.invoice_date AS "Belegdatum",
            p.pos_number AS "Position",
            p.gesamt_netto AS "Positions-Netto",
            p.gesamtpreis AS "Positions-Brutto"
        FROM invoice_pos p
        JOIN invoices i ON i.id = p.invoice_id
        JOIN clients c ON c.id = i.client_id
        ORDER BY p.invoice_id, p.pos_number
    """


def _format_german_decimal(value: Any) -> str:
    if value is None:
        return ""
    amount = Decimal(str(value)).quantize(Decimal("0.01"))
    return f"{amount:.2f}".replace(".", ",")


def _format_export_value(value: Any, column: str) -> Any:
    if value is None:
        return ""
    if column in {"Belegdatum"}:
        return _format_date(value)
    if column in {"Netto", "USt", "Brutto", "Positions-Netto", "Positions-Brutto"}:
        return _format_german_decimal(value)
    return value


def _format_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    try:
        parsed = datetime.fromisoformat(str(value))
    except ValueError:
        return str(value)
    return parsed.strftime("%d.%m.%Y")


def _write_row(
    worksheet,
    row: int,
    start_column: int,
    values: list[Any],
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    border: Border | None = None,
    alignment: Alignment | None = None,
) -> None:
    for offset, value in enumerate(values):
        cell = worksheet.cell(row=row, column=start_column + offset, value=value)
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        if border is not None:
            cell.border = border
        if alignment is not None:
            cell.alignment = alignment


def _fit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 44)
